import openpyxl as xl
from openpyxl.chart import bar_chart, reference
wb = xl.load_workbook('transection.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)

for row in range(1, sheet.max_row +1):
    print(row)
print("--------------")
for row in range(2, sheet.max_row +1):
    cell = sheet.cell(row, 3)
    print(cell.value)

print("---------------")
for row in range(2, sheet.max_row +1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell = corrected_price = corrected_price   


wb.save('transection2.xlsx')