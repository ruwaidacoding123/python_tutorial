import openpyxl as xl
from openpyxl.chart import bar_chart, reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']


    for row in range(1, sheet.max_row +1):
        print(row)
    print("--------------")
    for row in range(2, sheet.max_row +1):
        cell = sheet.cell(row, 3)
        print(cell.value)

    print("---------------")
    for row in range(2, sheet.max_row +1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell = corrected_price = corrected_price

        value = Reference(sheet,
                          min_row=2,
                          max_row=sheet.max_row,
                          min_col=4,
                          max_col=4)
        chart = bar_chart()
        chart.add_data(values)
        sheet.add_chart(chart, 'e2')


    wb.save(filename)